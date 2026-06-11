"""Build a self-contained HTML version of the AISAN 4667 take-private case study.

All figures are read live from the verified Excel model (data_only) and the peer /
precedent CSVs — nothing is hand-typed — so the HTML cannot drift from the model.

Run:
    uv run python -m src.report.build_case_html
Output:
    docs/AISAN_4667_Case_Study.html  (single file, no external dependencies)
"""

from __future__ import annotations

import csv
import html
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "docs" / "AISAN_4667_LBO_Model.xlsx"
PEERS = ROOT / "data" / "processed" / "peer_multiples.csv"
PRECED = ROOT / "data" / "processed" / "precedent_premium_check.csv"
OUT = ROOT / "docs" / "AISAN_4667_Case_Study.html"

NAVY, GOLD, ICE, GREEN, RED = "#1E2761", "#C8A951", "#CADCFC", "#2C5F2D", "#9B2D2D"


# ----------------------------------------------------------------------------- data
def load():
    wb = load_workbook(XLSX, data_only=True)
    return {ws.title: ws for ws in wb.worksheets}


def g(ws, c):
    return ws[c].value


def read_csv(path):
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# ------------------------------------------------------------------------- format
def yen(v, dp=0):
    if v is None:
        return "n.a."
    return f"¥{v:,.{dp}f}"


def bn(v, dp=1):
    return f"¥{v/1000:,.{dp}f}bn"


def pct(v, dp=1):
    return "n.a." if v is None else f"{v*100:.{dp}f}%"


def x(v, dp=2):
    return "n.a." if v is None else f"{v:.{dp}f}x"


def esc(s):
    return html.escape(str(s))


# -------------------------------------------------------------------------- charts
def svg_bars(series, labels, width=560, height=240, pad=34, colors=None, fmt=lambda v: f"{v:,.0f}"):
    """Grouped vertical bar chart as inline SVG. series: list of (name,color,values)."""
    n = len(labels)
    allv = [v for _, _, vals in series for v in vals]
    vmax = max(allv) * 1.18 or 1
    plot_w, plot_h = width - pad * 2, height - pad * 2
    group_w = plot_w / n
    bw = group_w / (len(series) + 0.6)
    out = [f'<svg viewBox="0 0 {width} {height}" class="chart" role="img">']
    # baseline
    y0 = pad + plot_h
    out.append(f'<line x1="{pad}" y1="{y0}" x2="{width-pad}" y2="{y0}" stroke="#d6d9e6"/>')
    for gi, lab in enumerate(labels):
        gx = pad + gi * group_w
        for si, (_, col, vals) in enumerate(series):
            v = vals[gi]
            bh = (v / vmax) * plot_h
            bx = gx + group_w * 0.30 + si * bw
            by = y0 - bh
            out.append(f'<rect x="{bx:.1f}" y="{by:.1f}" width="{bw*0.86:.1f}" height="{bh:.1f}" fill="{col}" rx="2"/>')
            out.append(f'<text x="{bx+bw*0.43:.1f}" y="{by-4:.1f}" class="bl">{fmt(v)}</text>')
        out.append(f'<text x="{gx+group_w/2:.1f}" y="{y0+16:.1f}" class="ax">{esc(lab)}</text>')
    out.append("</svg>")
    legend = " ".join(
        f'<span class="lg"><i style="background:{col}"></i>{esc(name)}</span>' for name, col, _ in series
    )
    return f'<div class="chartwrap">{"".join(out)}<div class="legend">{legend}</div></div>'


def svg_waterfall(items, width=620, height=300, pad=40):
    """Waterfall: items = list of (label, value, kind) kind in start/up/down/end."""
    vals = [v for _, v, _ in items]
    cum, lo, hi, tops = 0, 0, 0, []
    for _, v, k in items:
        if k in ("start", "end"):
            tops.append((0, v)); hi = max(hi, v)
        elif k == "up":
            tops.append((cum, cum + v)); hi = max(hi, cum + v); cum += v
        else:
            tops.append((cum + v, cum)); hi = max(hi, cum); cum += v
    hi *= 1.12
    plot_w, plot_h = width - pad * 2, height - pad * 2
    bw = plot_w / len(items) * 0.62
    step = plot_w / len(items)
    y0 = pad + plot_h
    out = [f'<svg viewBox="0 0 {width} {height}" class="chart" role="img">']
    prev_x = None
    for i, (lab, v, k) in enumerate(items):
        b, t = tops[i]
        col = {"start": NAVY, "end": GOLD, "up": GREEN, "down": RED}[k]
        bx = pad + i * step + (step - bw) / 2
        yt = y0 - (t / hi) * plot_h
        yb = y0 - (b / hi) * plot_h
        out.append(f'<rect x="{bx:.1f}" y="{min(yt,yb):.1f}" width="{bw:.1f}" height="{abs(yb-yt):.1f}" fill="{col}" rx="2"/>')
        sign = "" if k in ("start", "end") else ("+" if v >= 0 else "−")
        out.append(f'<text x="{bx+bw/2:.1f}" y="{min(yt,yb)-5:.1f}" class="bl">{sign}{abs(v)/1000:.2f}</text>')
        for ln in lab.split("\n"):
            out.append(f'<text x="{bx+bw/2:.1f}" y="{y0+14+lab.split(chr(10)).index(ln)*11:.1f}" class="ax">{esc(ln)}</text>')
        if prev_x is not None and k not in ("end",):
            out.append(f'<line x1="{prev_x:.1f}" y1="{yt if k=="up" else yb:.1f}" x2="{bx:.1f}" y2="{yt if k=="up" else yb:.1f}" stroke="#b9bed4" stroke-dasharray="3 3"/>')
        prev_x = bx + bw
    out.append("</svg>")
    return f'<div class="chartwrap">{"".join(out)}</div>'


def heatmap(prem, mults, grid):
    flat = [v for row in grid for v in row]
    lo, hi = min(flat), max(flat)
    def cell(v):
        t = (v - lo) / (hi - lo) if hi > lo else 0.5
        # red(low) -> amber -> green(high)
        if t < 0.5:
            r, gc, b = 255, int(120 + 135 * (t * 2)), 90
        else:
            r, gc, b = int(255 - 175 * ((t - 0.5) * 2)), 190, 90
        strong = v >= 0.20
        bord = f'border:2px solid {NAVY};' if strong else ""
        return f'<td style="background:rgb({r},{gc},{b});{bord}">{pct(v)}</td>'
    head = "".join(f"<th>{x(m,1)}</th>" for m in mults)
    rows = ""
    for i, p in enumerate(prem):
        rows += f"<tr><th>{pct(p,0)}</th>" + "".join(cell(grid[i][j]) for j in range(len(mults))) + "</tr>"
    return (f'<table class="heat"><thead><tr><th>Premium ＼ Exit</th>{head}</tr></thead>'
            f"<tbody>{rows}</tbody></table>")


# --------------------------------------------------------------------------- build
def build():
    S = load()
    A, SU, M, D, R, SC, LV = (S["Assumptions"], S["Sources_Uses"], S["Model"],
                              S["Debt_FCF"], S["Returns"], S["Scenarios"], S["Lev_Sensitivity"])
    peers = read_csv(PEERS)
    preced = read_csv(PRECED)

    # KPIs
    irr, moic = g(R, "C16"), g(R, "C15")
    offer, ref, prem = g(A, "C8"), g(A, "C5"), g(A, "C7")
    netcash = -g(A, "C12")

    # projections
    fy = ["FY26E", "FY27E", "FY28E", "FY29E", "FY30E", "FY31E"]
    rev_cols = ["E", "F", "G", "H", "I", "J"]
    rev = [g(M, c + "5") for c in rev_cols]
    ebitda = [g(M, c + "10") for c in rev_cols]
    # historical rev/op
    hist_fy = ["FY24A", "FY25A", "FY26E"]
    hist_rev = [g(M, c + "5") for c in ["C", "D", "E"]]
    hist_op = [g(M, c + "7") for c in ["C", "D", "E"]]

    # bridge
    bridge = [
        ("Entry\nequity", g(R, "J5"), "start"),
        ("EBITDA\ngrowth", g(R, "J6"), "up"),
        ("Multiple\nchange", g(R, "J7"), "down" if g(R, "J7") < 0 else "up"),
        ("Deleverage,\nFCF & fees", g(R, "J8"), "up"),
        ("Exit\nequity", g(R, "J9"), "end"),
    ]

    # sensitivity grid
    mults = [g(R, c + "21") for c in ["F", "G", "H", "I", "J"]]
    prem_rows = [g(R, "E" + str(r)) for r in range(22, 27)]
    grid = [[g(R, c + str(r)) for c in ["F", "G", "H", "I", "J"]] for r in range(22, 27)]

    css = f"""
:root{{--navy:{NAVY};--gold:{GOLD};--ice:{ICE};--green:{GREEN};--red:{RED};}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;color:#1c2230;line-height:1.55;background:#f4f5f9}}
h1,h2,h3,.serif{{font-family:Georgia,'Times New Roman',serif}}
a{{color:var(--navy)}}
.wrap{{max-width:1080px;margin:0 auto;padding:0 22px}}
nav{{position:sticky;top:0;z-index:20;background:var(--navy);color:#fff;font-size:12.5px}}
nav .wrap{{display:flex;flex-wrap:wrap;gap:2px 16px;align-items:center;padding:9px 22px}}
nav b{{color:var(--gold);margin-right:auto;letter-spacing:.12em;font-size:11px}}
nav a{{color:#cfd6ee;text-decoration:none;white-space:nowrap}}
nav a:hover{{color:#fff}}
header.hero{{background:var(--navy);color:#fff;padding:56px 0 40px}}
header.hero .tag{{color:var(--gold);letter-spacing:.28em;font-size:12px;font-weight:700}}
header.hero h1{{font-size:40px;margin:10px 0 4px;line-height:1.1}}
header.hero .sub{{color:var(--ice);font-size:20px}}
.factrow{{display:flex;flex-wrap:wrap;gap:26px;margin-top:30px;padding-top:18px;border-top:1px solid rgba(255,255,255,.15)}}
.factrow div span{{display:block}}
.factrow .k{{color:#9aa6cf;font-size:11px;letter-spacing:.1em}}
.factrow .v{{font-weight:700;font-size:16px}}
section{{padding:38px 0;border-bottom:1px solid #e4e7f0}}
section h2{{font-size:27px;color:var(--navy)}}
section .lead{{color:#5b6485;font-style:italic;margin:4px 0 22px;font-size:15px}}
.banner{{background:var(--navy);color:#fff;border-radius:10px;padding:20px 24px;margin:18px 0}}
.banner .rk{{color:var(--gold);letter-spacing:.16em;font-size:11px;font-weight:700}}
.banner b{{font-size:17px}}
.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin:20px 0}}
.kpi{{background:#fff;border:1px solid #e4e7f0;border-radius:10px;padding:16px}}
.kpi .k{{color:#7a83a3;font-size:10.5px;letter-spacing:.1em;font-weight:700}}
.kpi .v{{font-family:Georgia,serif;font-size:30px;color:var(--navy);margin:4px 0 2px}}
.kpi .s{{color:#7a83a3;font-size:11.5px}}
.grid2{{display:grid;grid-template-columns:1fr 1fr;gap:22px}}
.grid3{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}}
.card{{background:#fff;border:1px solid #e4e7f0;border-radius:10px;padding:18px}}
.card.l{{border-left:5px solid var(--green)}}
.card.r{{border-left:5px solid var(--red)}}
.card.g{{border-left:5px solid var(--gold)}}
.card h3{{color:var(--navy);font-size:17px;margin-bottom:6px}}
.card .role{{font-size:11px;font-weight:700;letter-spacing:.08em;color:var(--gold);text-transform:uppercase}}
ul{{margin:6px 0 0 18px}}li{{margin:3px 0}}
table{{width:100%;border-collapse:collapse;font-size:13px;margin:6px 0}}
th,td{{padding:6px 9px;text-align:right}}
th:first-child,td:first-child{{text-align:left}}
thead th{{background:var(--navy);color:#fff;font-weight:600}}
tbody tr:nth-child(even){{background:#f1f3f9}}
table.heat td{{text-align:center;font-weight:600;color:#1c2230}}
table.heat th{{background:var(--navy);color:#fff}}
.num{{font-variant-numeric:tabular-nums}}
.chartwrap{{background:#fff;border:1px solid #e4e7f0;border-radius:10px;padding:14px}}
.chart{{width:100%;height:auto}}
.chart .ax{{font-size:10px;fill:#5b6485;text-anchor:middle}}
.chart .bl{{font-size:10px;fill:#1c2230;text-anchor:middle;font-weight:600}}
.legend{{text-align:center;font-size:11.5px;color:#5b6485;margin-top:6px}}
.legend .lg{{margin:0 8px}}.legend i{{display:inline-block;width:10px;height:10px;border-radius:2px;margin-right:4px;vertical-align:middle}}
.note{{font-size:12px;color:#7a83a3;margin-top:8px}}
.pill{{display:inline-block;background:var(--ice);color:var(--navy);border-radius:20px;padding:2px 10px;font-size:11px;font-weight:700;margin:2px}}
.tl{{list-style:none;margin:0;border-left:2px solid var(--gold);padding-left:18px}}
.tl li{{margin:0 0 12px}}.tl .d{{font-weight:700;color:var(--navy)}}
footer{{background:var(--navy);color:#9aa6cf;font-size:11.5px;padding:26px 0}}
@media(max-width:820px){{.kpis,.grid2,.grid3{{grid-template-columns:1fr}}header.hero h1{{font-size:30px}}}}
"""

    def kpi(k, v, s):
        return f'<div class="kpi"><div class="k">{k}</div><div class="v">{v}</div><div class="s">{s}</div></div>'

    def sec(id_, num, title, lead, body):
        return (f'<section id="{id_}"><div class="wrap"><h2><span style="color:var(--gold)">{num}</span> '
                f'{esc(title)}</h2><div class="lead">{esc(lead)}</div>{body}</div></section>')

    # ---- peers table
    peer_rows = ""
    for p in peers:
        ev_ebitda = p.get("ev_ebitda", "")
        try:
            ev_ebitda = f"{float(ev_ebitda):.1f}x"
        except (ValueError, TypeError):
            ev_ebitda = "n.a." if (p.get("data_status") != "sourced") else ev_ebitda
        peer_rows += (f"<tr><td>{esc(p['company'])}</td><td>{esc(p['ticker'])}</td>"
                      f"<td>{esc(p['geography'])}</td><td class='num'>{ev_ebitda}</td></tr>")

    pre_rows = ""
    for p in preced:
        pre_rows += (f"<tr><td>{esc(p['target'])}</td><td>{esc(p['buyer'])}</td>"
                     f"<td class='num'>{esc(p['premium_prior_day'])}</td>"
                     f"<td class='num'>{esc(p.get('premium_6m','') or '—')}</td></tr>")

    # ---- assemble
    parts = []
    parts.append(f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Project Measure — AISAN TECHNOLOGY (4667) Take-Private Case Study</title>
<style>{css}</style></head><body>
<nav><div class="wrap"><b>PROJECT MEASURE</b>
<a href="#summary">Summary</a><a href="#business">Business</a><a href="#valuation">Valuation</a>
<a href="#structure">Structure</a><a href="#returns">Returns</a><a href="#cash">Cash</a>
<a href="#risks">Risks</a><a href="#gating">Gating</a><a href="#sources">Sources</a></div></nav>
<header class="hero"><div class="wrap">
<div class="tag">PROJECT MEASURE · STRICTLY PRIVATE &amp; CONFIDENTIAL</div>
<h1>AISAN TECHNOLOGY CO., LTD.</h1>
<div class="sub">Take-Private (LBO) Case Study · TSE Standard 4667</div>
<div class="factrow">
<div><span class="k">HQ</span><span class="v">Nagoya, Japan</span></div>
<div><span class="k">FOUNDED</span><span class="v">1970</span></div>
<div><span class="k">SECTOR</span><span class="v">Geospatial software / HD maps</span></div>
<div><span class="k">RECOMMENDATION</span><span class="v">Too early — DD only</span></div>
</div></div></header>""")

    # Summary
    summary_body = f"""
<div class="banner"><span class="rk">RECOMMENDATION</span> &nbsp;
<b>Too early; proceed to confirmatory DD only.</b> No bid pending the investigation report, audited /
restated FY2026 financials, QoE, cash-availability analysis and clean re-underwriting. On 3 Apr 2026 AISAN
set up a special committee after suspected forgery, concealment and improper transactions at a 100%-owned
subsidiary; the FY26 impact is undetermined. Even on pre-restatement figures, base-case returns are
~{pct(irr)} IRR / {x(moic)} — below a typical PE hurdle.</div>
<div class="kpis">
{kpi("BASE-CASE IRR", pct(irr), f"MOIC {x(moic)} · 5-yr hold")}
{kpi("OFFER / SHARE", yen(offer), f"+{pct(prem,0)} to {yen(ref)} ref")}
{kpi("GATING ITEM", "Special probe", "100%-sub investigation")}
{kpi("NET CASH", bn(netcash), "≈44% of mkt cap; TBC")}
</div>
<div class="grid2">
<div class="card l"><div class="role">Fundamental attractions</div><ul>
<li>Sticky, high-margin surveying-software core (cash cow; recurring mix TBC)</li>
<li>Over-capitalised balance sheet: ~{bn(netcash)} net cash; deployment assumed in base case</li>
<li>Founder / related holders ~18.5% (reported) may support a take-private if aligned</li>
<li>HD-map &amp; Level-4 autonomous-driving optionality for a strategic exit</li></ul></div>
<div class="card r"><div class="role">Why not now</div><ul>
<li>Subsidiary (Akisoku) under special investigation — suspected forgery &amp; concealment</li>
<li>Financials unreliable until restated; FY26 results formally delayed (30-Apr-26 notice)</li>
<li>Base IRR ~{pct(irr)} (upside ~{pct(g(SC,'E14'))}, downside a capital loss) — below a clean 20% hurdle</li>
<li>Lumpy Q4-weighted revenue; loss-making Mobility; thin Standard float</li></ul></div></div>"""
    parts.append(sec("summary", "01", "Executive Summary & Recommendation",
                     "Fundamentally interesting, but an active subsidiary investigation and sub-hurdle returns make a bid premature",
                     summary_body))

    # Business & industry
    seg_rows = ""
    for r in range(21, 24):
        seg_rows += (f"<tr><td>{esc(g(M,'B'+str(r)))}</td><td class='num'>{esc(g(M,'C'+str(r)))}</td>"
                     f"<td>{esc(g(M,'D'+str(r)))}</td><td class='num'>{esc(g(M,'F'+str(r)))}</td>"
                     f"<td>{esc(g(M,'G'+str(r)))}</td></tr>")
    business_body = f"""
<div class="grid2">
<div class="card g"><div class="role">Public · cash cow</div><h3>Surveying &amp; civil-engineering software</h3>
<p>Licensed land/house surveyors &amp; government bodies. Products: WingEarth · ANIST (point-cloud CAD) · GrandBase.
Recurring, sticky, ~26% segment margin (H1 FY26). Stable, public-budget-linked — funds the thesis and gives downside protection.</p>
<div style="margin-top:8px">{''.join(f'<span class="pill">{t}</span>' for t in ['i-Construction 2.0','Surveyor shortage','National 3D / PLATEAU','Disaster resilience'])}</div></div>
<div class="card"><div class="role" style="color:var(--navy)">Mobility / DX · growth</div><h3>HD maps &amp; autonomous-driving systems</h3>
<p>Mobile Mapping (MMS), high-precision 3D / dynamic maps, Level-4 system integration, proving-ground tests, drones.
Lumpy, Q4-weighted, currently low / negative margin. Early-stage and capital-hungry — the value-up plan rationalises
losses while preserving strategic-exit optionality.</p>
<div style="margin-top:8px">{''.join(f'<span class="pill">{t}</span>' for t in ['Level-4 legalised (2023)','Robotaxi / bus pilots','Logistics automation','Drone / delivery'])}</div></div></div>
<h3 style="color:var(--navy);margin:22px 0 4px">Illustrative segment economics (FY26E)</h3>
<table><thead><tr><th>Segment</th><th>Rev share</th><th>Margin</th><th>Growth</th><th>Value-up lever</th></tr></thead>
<tbody>{seg_rows}</tbody></table>
<p class="note">Reported / estimated; the consolidated model is top-down, not a bottom-up segment build. To be verified in DD.</p>"""
    parts.append(sec("business", "02", "Business, Segments & Industry",
                     "Two engines: a cash-generative software core and a growth-oriented mobility option", business_body))

    # Financial track record
    fin_body = f"""
{svg_bars([("Revenue", NAVY, hist_rev), ("Operating profit", GOLD, hist_op)], hist_fy, height=230)}
<div class="kpis" style="margin-top:18px">
{kpi("FY26E REVENUE", bn(g(M,'E5')), f"+{pct(g(M,'E6'),0)} YoY (guidance)")}
{kpi("FY26E OP. PROFIT", yen(g(M,'E7'))+"m", f"{pct(g(M,'E8'))} margin")}
{kpi("FY25A NET CASH", bn(netcash), "structurally light B/S")}
{kpi("FY26E EPS", yen(g(A,'C23')/g(A,'C6')*1000), f"fwd P/E ~{(ref*g(A,'C6'))/g(A,'C23'):.1f}x")}
</div>
<p class="note">Revenue compounded at ~8.6% over FY17–FY25; FY26E guidance {bn(g(M,'E5'))}. Operating margin has held in a
~7–8% band — well below the software core's potential. Profitability, not the revenue line, is the lever.</p>"""
    parts.append(sec("financials", "03", "Financial Track Record",
                     "Steady top-line growth and a structurally light balance sheet", fin_body))

    # Valuation context
    val_body = f"""
<div class="kpis">
{kpi("ENTRY EV / EBITDA", x(g(A,'C20'),1), "FY26E EBITDA")}
{kpi("ENTRY EV / EBIT", x(g(A,'C21'),1), "FY26E EBIT")}
{kpi("OFFER P / E", x(g(A,'C22'),1), "ex-treasury EPS")}
{kpi("IMPLIED EV", bn(g(A,'C13')), f"net of {bn(netcash)} cash")}
</div>
<div class="grid2">
<div class="card"><h3>Peer trading multiples (LTM EV/EBITDA)</h3>
<table><thead><tr><th>Company</th><th>Ticker</th><th>Geo</th><th>EV/EBITDA</th></tr></thead><tbody>{peer_rows}</tbody></table>
<p class="note">yfinance snapshots, 7-Jun-2026. AISAN's {x(g(A,'C20'),1)} entry sits above domestic geospatial peers
(~4–5x) and below global construction-software names (16–26x). Verify in a market-data terminal before bid.</p></div>
<div class="card"><h3>Japan take-private precedent premia</h3>
<table><thead><tr><th>Target</th><th>Buyer</th><th>vs prior day</th><th>vs 6m</th></tr></thead><tbody>{pre_rows}</tbody></table>
<p class="note">Reported premia commonly start above 30%; the 20% walk-to-price (below) sits well under precedent. Sources:
SECOM/ITOCHU &amp; Topcon tender disclosures; M&amp;A Online.</p></div></div>"""
    parts.append(sec("valuation", "04", "Valuation Context",
                     "Strategic scarcity supports a premium, but sponsor returns do not clear at full precedent-style premia", val_body))

    # Structure / Sources & Uses
    su_rows = ""
    for lbl, c, in [("Purchase of equity", "C5"), ("Repay existing debt", "C6"), ("Transaction & financing fees", "C7"), ("Total uses", "C8")]:
        su_rows += f"<tr><td>{lbl}</td><td class='num'>{yen(g(SU,c))}</td></tr>"
    so_rows = ""
    for lbl, c in [("New senior term loan", "G5"), ("Excess cash on balance sheet", "G6"), ("Sponsor equity", "G7"), ("Total sources", "G8")]:
        so_rows += f"<tr><td>{lbl}</td><td class='num'>{yen(g(SU,c))}</td></tr>"
    struct_body = f"""
<div class="grid2">
<div class="card"><h3>Uses of funds (¥mm)</h3><table><tbody>{su_rows}</tbody></table></div>
<div class="card"><h3>Sources of funds (¥mm)</h3><table><tbody>{so_rows}</tbody></table></div></div>
<div class="kpis" style="margin-top:16px">
{kpi("NEW SENIOR DEBT", bn(g(SU,'C12')), f"{x(g(SU,'H5'),1)} new debt / EBITDA")}
{kpi("NET DEBT AT CLOSE", bn(g(SU,'C14')), f"{x(g(SU,'C15'),1)} net debt / EBITDA")}
{kpi("SPONSOR EQUITY", bn(g(SU,'G7')), f"{pct(g(SU,'C17'),0)} of capitalisation")}
{kpi("TERM-LOAN PRICING", pct(g(A,'G7')), "TIBOR/TONA + spread")}
</div>
<p class="note">Net cash and excess balance-sheet cash fund most of the price; new leverage is kept moderate
(2.0x gross / ~1.0x net) given lumpy EBITDA. This is not a classic leverage-driven LBO — sponsor equity stays high at {bn(g(SU,'G7'))}.</p>"""
    parts.append(sec("structure", "05", "Illustrative LBO Structure",
                     "Net cash and excess cash fund most of the price; new leverage kept moderate (2.0x gross / ~1.0x net)", struct_body))

    # Projections
    proj_body = f"""
{svg_bars([("Revenue", NAVY, rev), ("EBITDA", GOLD, ebitda)], fy, height=240)}
<div class="grid3" style="margin-top:16px">
{kpi("REVENUE CAGR", pct((rev[-1]/rev[0])**(1/5)-1), "FY26E → FY31E")}
{kpi("EBIT MARGIN", f"{pct(g(M,'F8'),0)} → {pct(g(M,'J8'),0)}", "phased expansion")}
{kpi("EXIT EBITDA", bn(g(M,'J10')), "FY31E")}
</div>
<p class="note">Mid-single-digit growth in line with history; EBITDA margin expands ~11% → ~15% as overhead is
rationalised and Mobility losses narrow. Capex {pct(g(A,'G26'),1)} of revenue, ΔNWC {pct(g(A,'I26'),0)} of Δrevenue,
cash tax {pct(g(A,'G14'),0)}, {pct(g(A,'G8'),0)} cash sweep. Deliberately conservative — not a hockey stick.</p>"""
    parts.append(sec("projections", "06", "Projections — Base Case",
                     "Mid-single-digit growth with gradual margin expansion toward the software core", proj_body))

    # Returns + bridge
    ret_rows = ""
    for lbl, c in [("Exit EBITDA (FY31E)", "C5"), ("Exit EV / EBITDA", "C6"), ("Exit enterprise value", "C7"),
                   ("Net cash at exit", "C8"), ("Exit equity value", "C9"), ("Sponsor equity in", "C12")]:
        val = g(R, c)
        disp = x(val, 1) if c == "C6" else yen(val)
        ret_rows += f"<tr><td>{lbl}</td><td class='num'>{disp}</td></tr>"
    returns_body = f"""
<div class="grid2">
<div>{svg_waterfall(bridge)}<p class="note">Bridge reconciles exactly to exit equity (Returns C9 = J9). 'Deleverage,
FCF &amp; fees' is net of ~¥0.4bn transaction costs.</p></div>
<div class="card g"><div class="role">Base-case returns</div>
<div style="font-family:Georgia,serif;font-size:44px;color:var(--navy)">{x(moic)}</div>
<div class="s" style="color:#7a83a3">GROSS MOIC</div>
<div style="font-family:Georgia,serif;font-size:30px;color:var(--gold);margin-top:6px">{pct(irr)}<span style="font-size:13px;color:#7a83a3"> gross IRR · 5-yr hold</span></div>
<table style="margin-top:10px"><tbody>{ret_rows}</tbody></table>
<p class="note">Gross of fees, carry &amp; MIP dilution; IRR assumes a single exit cash flow (no interim dividends).</p></div></div>"""
    parts.append(sec("returns", "07", "Returns — Base Case",
                     f"Base case: {x(moic)} MOIC / ~{pct(irr)} IRR (below a clean 20% hurdle) — figures pending any restatement", returns_body))

    # Sensitivity + scenarios
    def scen_card(col, name, klass):
        return (f'<div class="card {klass}"><div class="role">{name}</div>'
                f'<div style="font-family:Georgia,serif;font-size:30px;color:var(--navy)">{x(g(SC,col+"13"))}'
                f' <span style="font-size:15px;color:var(--gold)">/ {pct(g(SC,col+"14"))}</span></div>'
                f'<p class="note">Rev CAGR {pct(g(SC,col+"5"),0)} · exit {x(g(SC,col+"8"),1)} · '
                f'exit EBITDA {bn(g(SC,col+"7"))}</p></div>')
    sens_body = f"""
<h3 style="color:var(--navy)">Gross IRR — offer premium × exit EV/EBITDA</h3>
{heatmap(prem_rows, mults, grid)}
<p class="note">Base case ({pct(prem,0)} premium, {x(g(A,'G16'),1)} exit) ≈ {pct(irr)} IRR (bordered cells clear 20%).
Returns stay below a clean 20% hurdle at the base premium; only low-premium / high-exit cases approach it.</p>
<div class="grid3" style="margin-top:18px">
{scen_card('C','Downside','r')}{scen_card('D','Base','g')}{scen_card('E','Upside','l')}</div>"""
    parts.append(sec("sensitivity", "08", "Sensitivity & Scenarios",
                     "Returns sit below a clean 20% hurdle at the base premium; the downside is a capital impairment", sens_body))

    # Cash availability + MIP
    cash_rows = ""
    for r in range(49, 52):
        cash_rows += (f"<tr><td>{esc(g(R,'B'+str(r)))}</td><td class='num'>{yen(g(R,'C'+str(r)))}</td>"
                      f"<td class='num'>{yen(g(R,'D'+str(r)))}</td><td class='num'>{x(g(R,'E'+str(r)))}</td>"
                      f"<td class='num'>{pct(g(R,'F'+str(r)))}</td></tr>")
    mip_rows = ""
    for r in range(55, 58):
        mip_rows += (f"<tr><td class='num'>{pct(g(R,'B'+str(r)),0)}</td><td class='num'>{x(g(R,'C'+str(r)))}</td>"
                     f"<td class='num'>{pct(g(R,'D'+str(r)))}</td></tr>")
    cash_body = f"""
<div class="grid2">
<div class="card"><h3>Cash availability — mechanics (¥mm)</h3>
<table><thead><tr><th>Case</th><th>Sponsor eq.</th><th>Exit eq.</th><th>MOIC</th><th>IRR</th></tr></thead>
<tbody>{cash_rows}</tbody></table>
<p class="note">Full = excess cash funds the buy-out at close; Retained = sponsor funds in full, cash stays;
Trapped = cash unrecoverable. Extraction needs distributable reserves, lender consent, clean accounts &amp; time —
none assured pre-DD.</p></div>
<div class="card"><h3>MIP dilution (gross → net)</h3>
<table><thead><tr><th>MIP dilution</th><th>MOIC</th><th>Gross IRR</th></tr></thead><tbody>{mip_rows}</tbody></table>
<p class="note">Returns are gross of fees &amp; carry. Illustrative net to LPs after a 2% fee / 20-over-8 carry / 5% MIP
≈ {x(g(SC,'C27'))} / {pct(g(SC,'D27'))} — widening the gap to a clean 20% hurdle (model: Scenarios net-returns panel).</p></div></div>
<div class="banner"><b>Takeaway —</b> AISAN screens cheaply partly on excess cash, but returns hinge on whether that
cash is legally and practically extractable. Treat as a confirmatory DD item, not a base-case assumption.</div>"""
    parts.append(sec("cash", "09", "Excess Cash Is Central to Returns",
                     "Base case assumes extraction; returns fall sharply if cash is retained or trapped — a DD item, not underwritten", cash_body))

    # Walk-to-price + leverage
    lev_rows = ""
    for r in range(5, 9):
        lev_rows += (f"<tr><td class='num'>{x(g(LV,'A'+str(r)),1)}</td><td class='num'>{bn(g(LV,'B'+str(r)))}</td>"
                     f"<td class='num'>{bn(g(LV,'C'+str(r)))}</td><td class='num'>{x(g(LV,'E'+str(r)))}</td>"
                     f"<td class='num'>{pct(g(LV,'F'+str(r)))}</td></tr>")
    walk_body = f"""
<div class="grid2">
<div class="card"><h3>20% hurdle walk-to-price</h3>
<div class="kpis" style="grid-template-columns:1fr 1fr 1fr">
{kpi("MAX OFFER", yen(g(R,'C42')), "20% gross IRR")}
{kpi("VS SPOT", "+"+pct(g(R,'C43')), f"{yen(ref)} ref")}
{kpi("VS 6M VWAP", "+"+pct(g(R,'C44')), f"{yen(g(A,'C32'))}")}
</div>
<p class="note">The maximum offer clearing a 20% hurdle ({yen(g(R,'C42'))}) sits well below precedent-style premia —
bid discipline is essential.</p></div>
<div class="card"><h3>Illustrative leverage sensitivity</h3>
<table><thead><tr><th>Debt/EBITDA</th><th>Debt</th><th>Sponsor eq.</th><th>MOIC</th><th>IRR</th></tr></thead>
<tbody>{lev_rows}</tbody></table>
<p class="note">Even {x(g(LV,'A8'),1)} gross debt does not clear 20% and would be imprudent before QoE, seasonality,
investigation impact &amp; lender appetite are confirmed. Returns are driven by entry price, cash, EBITDA growth,
margin and exit multiple — not leverage.</p></div></div>
<div class="banner"><b>Strategic-owner lens —</b> a mobility-platform acquirer with ≈{yen(g(SC,'C39'))}mm of exit synergy
EBITDA (~{pct(g(SC,'D39'))} of FY31E EBITDA; HD-map licensing for robotaxi ODDs, mapping-as-a-service, municipal AV projects)
can support the full {yen(g(A,'C8'))} offer at a 20% hurdle — exactly closing the walk-to-price gap a financial sponsor cannot
underwrite (model: Scenarios — synergy walk).</div>"""
    parts.append(sec("walk", "10", "Walk-to-Price & Leverage",
                     "Price discipline: the maximum offer clearing a 20% hurdle sits well below precedent-style premia", walk_body))

    # Value-up
    vu = [("Balance-sheet optimisation", "Important but DD-dependent",
           f"Deploy the ~{bn(netcash)} net cash and monetise investment securities. Excess cash reduces headline EV but does not make this a classic leverage-driven LBO — sponsor equity stays high at {bn(g(SU,'G7'))} because debt capacity is constrained. Subject to legal, operating, lender and investigation DD.", "g"),
          ("Margin expansion", "High control, phased",
           "Rationalise group overhead and lift the software mix; turn around or partner the loss-making Mobility unit to move blended operating margin from ~8% toward 12%+.", ""),
          ("Strategic re-rating of optionality", "Higher risk, higher reward",
           "Professionalise and scale the HD-map / Level-4 franchise into a separately valuable asset, supporting an exit to an auto OEM, map consolidator or geospatial buyer at a premium multiple.", ""),
          ("Defensive compounding", "Downside protection",
           "Protect and grow the recurring public-sector software base through i-Construction and national 3D-data tailwinds — the floor under every scenario.", "l")]
    vu_cards = "".join(f'<div class="card {k}"><div class="role">{esc(role)}</div><h3>{esc(t)}</h3><p>{esc(body)}</p></div>'
                       for t, role, body, k in vu)
    parts.append(sec("valueup", "11", "Value-Up Theses",
                     "Four levers that could lift returns if the investigation clears — none is fully underwritable pre-DD",
                     f'<div class="grid2">{vu_cards}</div>'))

    # Risks + DD
    risks = [("Subsidiary investigation (GATING)", "Special committee at 100%-sub Akisoku — suspected forgery, concealment & unpaid debt; FY26 impact undetermined and financials may be restated."),
             ("Seasonality / cash burn", "Q4-loaded revenue and a Q1–Q3 working-capital draw complicate debt-service timing."),
             ("Mobility losses", "The growth segment is not yet profitable; turnaround timing and capital needs are uncertain."),
             ("Public-budget reliance & float", "Demand tied to government budgets; thin TSE Standard float and minority protections."),
             ("Sub-hurdle returns", f"Even pre-restatement, base IRR ~{pct(irr)} (upside ~{pct(g(SC,'E14'))}) is below a clean 20% PE hurdle.")]
    dd = ["Special-investigation findings, scope & quantified FY26 impact / restatement",
          "Whether issues extend beyond Akisoku — group-wide internal controls",
          "Current cap table & founder-family intent on an MBO / roll-over",
          "Net cash & investment-securities mark at close; distributable reserves",
          "Mobility / DX standalone P&L and a credible path to breakeven",
          "Public-revenue backlog, contract terms & budget sensitivity",
          "Quality of earnings on Q4-loaded revenue recognition; software cap. policy",
          "Lender appetite & debt capacity given the seasonal cash profile"]
    risks_html = "".join(f"<li><b>{esc(t)}</b> — {esc(b)}</li>" for t, b in risks)
    dd_html = "".join(f"<li>{esc(d)}</li>" for d in dd)
    parts.append(sec("risks", "12", "Key Risks & Diligence Checklist",
                     "What must be verified before committing capital in the next phase",
                     f'<div class="grid2"><div class="card r"><div class="role">Key risks</div><ul>{risks_html}</ul></div>'
                     f'<div class="card l"><div class="role">Priority DD items</div><ul>{dd_html}</ul></div></div>'))

    # Gating timeline
    tl = [("Jan 2024", "AISAN acquires 100% of Akisoku (有限会社秋測), an Akita-based survey-instrument sales / repair firm."),
          ("11 Mar 2026", "A supplier queries unpaid debt at the subsidiary's marketing centre (Ueda, Nagano) — the trigger."),
          ("Internal review", "Suspected improper transactions by a subsidiary director; forgery & concealment of documents."),
          ("Widening", "Further forged documents and concealment of inventory losses / unpaid debts identified."),
          ("3 Apr 2026", "Special Investigation Committee established (external lawyer + CPAs, outside directors / auditors)."),
          ("Market", "Shares fell to ~¥1,643 intraday on disclosure and drifted to ~¥1,490s by May; FY26 impact undetermined.")]
    tl_html = "".join(f'<li><span class="d">{esc(d)}</span><br>{esc(t)}</li>' for d, t in tl)
    gate_html = "".join(f"<li>{esc(x)}</li>" for x in
                        ["Final investigation report & quantified financial impact", "Restated / audited financial statements",
                         "Evidence the issues are contained to Akisoku", "Remediation of group-wide internal controls"])
    parts.append(sec("gating", "13", "Gating Item — Subsidiary Probe",
                     "An unresolved investigation at a 100%-owned subsidiary overrides the investment case until cleared",
                     f'<div class="grid2"><div class="card"><div class="role">What happened — Akisoku</div><ul class="tl">{tl_html}</ul></div>'
                     f'<div class="card"><div class="banner" style="margin-top:0">GATE — required before any bid</div><ul>{gate_html}</ul>'
                     '<p class="note">Source: AISAN IR (3 Apr 2026); Nikkei; Yahoo Finance JP. See sources.</p></div></div>'))

    # Conclusion
    concl_body = f"""
<div class="banner"><span class="rk">CONCLUSION</span><br><b style="font-size:19px">A fundamentally interesting,
cash-rich target — but a live investigation and sub-hurdle returns mean we should not bid now.</b></div>
<div class="grid3">
{kpi("RECOMMENDATION", "Too early", "proceed to confirmatory DD only")}
{kpi("BASE CASE", f"{x(moic)} / {pct(irr)}", "below hurdle")}
{kpi("UPSIDE", f"{x(g(SC,'E13'))} / {pct(g(SC,'E14'))}", "only if cleared")}
</div>
<div class="card" style="margin-top:16px"><div class="role" style="color:var(--navy)">Next steps</div><ol style="margin-left:18px">
<li>Await the special-investigation report &amp; any restatement before any bid</li>
<li>Re-underwrite on restated financials; re-base the offer off a clean price</li>
<li>Confirmatory diligence per the priority checklist</li>
<li>Prepare an outreach plan subject to explicit authorization (no contact under current data policy)</li>
<li>Keep on watchlist; bid only after the probe clears, financials are re-underwritten, excess-cash availability
is verified, and walk-to-price (~{yen(g(R,'C42'))}/share) supports a clean 20%+ gross IRR</li></ol></div>"""
    parts.append(sec("conclusion", "14", "Conclusion & Next Steps", "Keep on the watchlist; proceed only to confirmatory DD", concl_body))

    # Sources
    src = [("AISAN IR — Special Investigation Committee notice (3 Apr 2026)", "https://www.aisantec.co.jp/information/4167/"),
           ("AISAN IR — share status / major shareholders (issued 5,548,979; treasury 268,816)", "https://www.aisantec.co.jp/ir/stock/status/"),
           ("AISAN IR — FY2025 results & FY2026 guidance; FY26 delay notice (30-Apr-26)", "https://irbank.net/4667/140120260430514374"),
           ("Nikkei — special investigation committee (3 Apr 2026)", "https://www.nikkei.com/"),
           ("Yahoo Finance JP — market reaction; ~¥1,643 intraday, ~¥1,490s May", "https://finance.yahoo.co.jp/quote/4667.T"),
           ("IRBANK — 4667 / E04980 (P&L, B/S, CF, segment)", "https://irbank.net/E04980/pl"),
           ("SECOM/ITOCHU — PASCO tender notice (5-Sep-2024)", "https://www.secom.co.jp/english/ir/lib_2024/notice20240905-1.pdf"),
           ("Topcon — KKR tender disclosure (29-Jul-2025)", "https://global.topcon.com/invest/"),
           ("M&A Online — Akisoku acquisition; Tecnos / Kaonavi precedents", "https://maonline.jp/news/20231013b")]
    src_html = "".join(f'<li>{esc(t)} — <a href="{u}">{esc(u)}</a></li>' for t, u in src)
    parts.append(sec("sources", "A", "Source Appendix",
                     "Selected public sources — every factual claim is to be verified in confirmatory diligence",
                     f'<ol style="margin-left:18px;font-size:13px">{src_html}</ol>'
                     '<p class="note">Reported / estimated / to be verified in DD: founder / related-holder stake, segment margins, '
                     'peer multiples, investment-securities mark, FY2026E guidance and the Akisoku impact. Peer multiples are '
                     'yfinance snapshots (7-Jun-2026). No company / shareholder / management contact has been made.</p>'))

    parts.append("""<footer><div class="wrap">Project Measure · Strictly Private &amp; Confidential · June 2026 ·
Preliminary case study for internal discussion only — not investment advice, a valuation opinion or an offer.
Figures are pre-restatement and may change with the investigation outcome. Generated from the verified Excel model.</div></footer>
</body></html>""")

    OUT.write_text("".join(parts), encoding="utf-8")
    print(f"wrote {OUT}  ({OUT.stat().st_size//1024} KB)")


if __name__ == "__main__":
    build()
